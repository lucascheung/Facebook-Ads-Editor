from facebook_business import FacebookAdsApi
from facebook_business.adobjects.adset import AdSet
from openpyxl import load_workbook
import deliveryestimate
import time
import os

print ('this is testing github')

my_app_id = '<INSERT APP AD>'
my_app_secret = '<INSERT APP SECRET'
my_access_token = '<INSERT ACCESS TOKEN>'
FacebookAdsApi.init(my_app_id, my_app_secret, my_access_token)

filepath = input("Input file path: ")
filepath = filepath.replace("'","")

filename = input("Insert Beat Name: ")

workbook = load_workbook(filepath)
sheet = workbook.active

script_fp = os.path.dirname(__file__)
templatefolder = r"/THE TEMPLATE/"
excelfile = "ReachEstimateTemplatev2.xlsx"
templatefilepath = "".join([script_fp,templatefolder,excelfile])

export_wb = load_workbook(templatefilepath)
export_ws = export_wb['Reach']

r=2
while sheet.cell(row=r,column=13).value != None:
    adset_id = str(sheet.cell(row=r,column = 13).value).replace("c:","")


    r+=1
    adset = AdSet(adset_id)
    
    #Pulls Adset Name
    AdsetDetail = adset.remote_read(fields=[AdSet.Field.name])
    AdsetDetailName = AdsetDetail["name"]


    #Pulls The Reach estimate for each adset
    reach_estimate = deliveryestimate.DeliveryEstimate.get_delivery_estimate(adset)
    TotalReach = reach_estimate[0]["estimate_mau"]
    

    export_ws.cell(row = r-1,column = 1).value = AdsetDetailName
    export_ws.cell(row = r-1,column = 2).value = int(TotalReach)
    export_ws.cell(row = r-1,column = 3).value = adset_id
    print ("Processed row {}".format(r-2))  

print ("\n\nProcessed all adsets!\n\n")

filetime = time.strftime("%Y%m%d-%H%M%S")
savefolder = r"/ReachEstimateFiles/"
excelfile = ".xlsx"
savingfilepath = "".join([script_fp,savefolder,filename,' - Facebook Reach Estimate ',filetime,excelfile])
export_wb.save(savingfilepath)
print('The Reach Estimate File is saved at ',savingfilepath)
try:  
    workbook.save(filepath)
except:
    print("Please close the file before running the script.")
